import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import math
import re
import io
import openai
import zipfile
from lxml import etree
from fpdf import FPDF

# 1. Page Setting
st.set_page_config(page_title="IS 2010 Scoring System", layout="wide")
st.title("IS 2010 Scoring System")

# --- Initialize OpenAI Client ---
try:
    client = openai.OpenAI(api_key=st.secrets["OPENAI_API_KEY"])

except Exception:
    client = None
    st.error("⚠️ OpenAI API Key not found. Please check your secrets.toml")

# --- Helper: Unified Formatting for Display ---
def format_ans(f, v):
    f_str = str(f).strip() if f is not None else ""
    v_str = str(v).strip() if v is not None else ""

    if f_str.startswith('='):
        return f"{f_str} ({v_str})"
    
    return v_str if v_str else "Empty"

# --- Core Grading Logic ---
def check_logic_equivalence(prof_f, stud_f, prof_v, stud_v):
    try:
        if prof_v is None and stud_v is None:
            return False
        
        v_p_float = float(prof_v) if prof_v is not None else 0.0
        v_s_float = float(stud_v) if stud_v is not None else 0.0
        values_match = math.isclose(v_p_float, v_s_float, rel_tol=1e-9, abs_tol=1e-9)

        
    except (ValueError, TypeError):
        values_match = str(prof_v).strip().upper() == str(stud_v).strip().upper()

    if values_match:
        p_is_formula = str(prof_f).startswith('=') if prof_f else False
        s_is_formula = str(stud_f).startswith('=') if stud_f else False
        
        if p_is_formula and s_is_formula:
            p_norm = str(prof_f).replace(" ", "").upper()
            s_norm = str(stud_f).replace(" ", "").upper()
            return p_norm == s_norm
        
        return True  # Pass if values match
        
    return False

def check_sparkline_advanced(p_cache, s_cache, cell_coord):

    def extract_xml_info(xml_cache, target_cell):
        try:
            clean_cell = target_cell.replace('$', '').upper()
            st.write(f"🔍 Finding {clean_cell}'s sparkline...")

            for xml_path, xml_content in xml_cache.items():
                st.write(f"Analyzing the file: {xml_path}")

                # Parse sparkline blocks using lxml instead of regex
                tree = etree.fromstring(xml_content.encode())
                NS = {
                    "x14": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main",
                    "xm":  "http://schemas.microsoft.com/office/excel/2006/main",
                }

                for sp in tree.findall('.//x14:sparkline', NS):
                    sq_elem = sp.find('xm:sqref', NS)
                    f_elem  = sp.find('xm:f',     NS)
                    if sq_elem is None or f_elem is None:
                        continue

                    norm_sq = sq_elem.text.replace('$', '').upper()

                    # Debug log
                    if clean_cell in norm_sq.split(' '):
                        st.write(f" [log] {clean_cell} cell found!")

                    if norm_sq == clean_cell:
                        # Traverse parent tags to read sparklineGroup attributes
                        grp = sp.getparent().getparent()  # sparkline → sparklines → sparklineGroup

                        found_range = f_elem.text.strip()
                        sp_type     = grp.get("type", "line")
                        markers     = grp.get("markers", "0") == "1"
                        high_point  = grp.get("highPoint", "0") == "1"
                        low_point   = grp.get("lowPoint", "0") == "1"
                        first_point = grp.get("firstPoint", "0") == "1"
                        last_point  = grp.get("lastPoint", "0") == "1"
                        negative    = grp.get("negative", "0") == "1"

                        st.write(f"🔎 [Sparkline Found] Cell: {clean_cell} -> Range: {found_range}")
                        return {
                            "range":       found_range,
                            "type":        sp_type,
                            "markers":     markers,
                            "high_point":  high_point,
                            "low_point":   low_point,
                            "first_point": first_point,
                            "last_point":  last_point,
                            "negative":    negative,
                        }

            return None

        except Exception as e:
            st.warning(f"[Error] {e}")
            return None

    p_info = extract_xml_info(p_cache, cell_coord)
    s_info = extract_xml_info(s_cache, cell_coord)

    if not p_info: return "Skip", None, None, None

    if not s_info: return False, "Missing", "Sparkline Object", "Sparkline is missing."

    p_src = p_info['range'].replace("$", "").replace(" ", "").upper()
    s_src = s_info['range'].replace("$", "").replace(" ", "").upper()

    if p_src != s_src:
        return False, f"Range: {s_info['range']}", f"Range: {p_info['range']}", "Data range error"

    if p_info['type'] != s_info['type']:
        return False, f"Type: {s_info['type']}", f"Type: {p_info['type']}", "Sparkline type error"

    marker_keys = ['markers', 'high_point', 'low_point', 'first_point', 'last_point', 'negative']
    if any(p_info[k] != s_info[k] for k in marker_keys):
        return False, "Marker settings differ", "See professor's file", "Marker configuration error"

    return True, None, None, None

# --- AI Feedback Function (called only when generating reports) ---
def get_ai_feedback(prof_ans_str, stud_ans_str, custom_msg=None):
    if not client: return "AI feedback disabled."
    if custom_msg: return custom_msg
    prompt = f"Compare Student Answer: {stud_ans_str} vs Professor's Correct Answer: {prof_ans_str}. Explain the error concisely under 150 words."

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "system", "content": "You are a professional Excel instructor."},
                      {"role": "user", "content": prompt}]
        )
        return response.choices[0].message.content
    except Exception as e: return f"AI Error: {e}"

def create_pdf_report(uid, score, errors):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(200, 15, txt="IS 2010 Lab Diagnostic Report", ln=True, align='C')
    pdf.set_font("Arial", '', 12)
    pdf.cell(200, 10, txt=f"Student ID: {uid} | Final Score: {score}", ln=True, align='C')
    pdf.line(10, 35, 200, 35)
    pdf.ln(10)
    if not errors:
        pdf.set_font("Arial", 'B', 14)
        pdf.cell(200, 10, txt="Congratulations! Perfect Score.", ln=True, align='C')
    else:
        for i, err in enumerate(errors):
            pdf.set_font("Arial", 'B', 11)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(0, 10, txt=f"{i+1}. Location: {err['Sheet']} Sheet / Cell {err['Cell']}", ln=True)
            
            pdf.set_font("Arial", '', 10)
            pdf.set_text_color(200, 0, 0)
            pdf.multi_cell(0, 7, txt=f"Your Answer: {err['Stud_Ans']}")
            pdf.set_text_color(0, 100, 0)
            pdf.multi_cell(0, 7, txt=f"Professor's Answer: {err['Prof_Ans']}")
            
            # Generate AI feedback only when needed
            feedback = get_ai_feedback(err['Prof_Ans'], err['Stud_Ans'], custom_msg=err.get('Custom_Msg'))
            
            pdf.set_fill_color(245, 245, 245)
            pdf.set_text_color(0, 50, 100)
            pdf.set_font("Arial", 'I', 10)

            feedback_str = feedback if feedback else "No feedback available."
            safe_feedback = feedback_str.encode('latin-1', 'replace').decode('latin-1')
            pdf.multi_cell(0, 8, txt=f"Feedback: {safe_feedback}", border=1, fill=True)
            pdf.ln(5)
    return pdf.output(dest='S').encode('latin-1', 'ignore')

# --- Initialize Session State ---
if 'grading_done' not in st.session_state:
    st.session_state.update({'grading_done': False, 'summary_df': pd.DataFrame(), 'all_wrongs_list': [], 'total_questions': 0})

with st.expander("Professor's Color Guide", expanded=False):
    standard_colors = {"Dark Red": "#C00000", "Red": "#FF0000", "Orange": "#FFC000", "Yellow": "#FFFF00", "Light Green": "#92D050", "Green": "#00B050", "Light Blue": "#00B0F0", "Blue": "#0070C0", "Dark Blue": "#002060", "Purple": "#7030A0"}
    cols = st.columns(10)
    st.caption("※ Only standard colors are supported for valid score calculations.")
    for idx, (name, hex_val) in enumerate(standard_colors.items()):
        cols[idx].markdown(f"<div style='background-color:{hex_val}; height:20px; border-radius:3px;'></div>", unsafe_allow_html=True)
        cols[idx].caption(name)

st.divider()

c1, c2 = st.columns(2)
prof_file = c1.file_uploader("1. Upload Professor's File", type=['xlsx'])
student_files = c2.file_uploader("2. Upload Student's File(s)", type=['xlsx'], accept_multiple_files=True)

if prof_file and student_files:
    selected_color = st.selectbox("Select Answer Cell Color:", ["--- Select ---"] + list(standard_colors.keys()))
    if st.button("Start Grading Process", width='stretch') and selected_color != "--- Select ---":
        target_hex = standard_colors[selected_color].lstrip('#').upper()
        p_bytes = prof_file.read()
        
        with zipfile.ZipFile(io.BytesIO(p_bytes)) as z:
            p_cache = {f: z.read(f).decode('utf-8', errors='replace') for f in z.namelist() if 'xl/worksheets/sheet' in f}
            
        p_wb_f = load_workbook(io.BytesIO(p_bytes), data_only=False, read_only=True)
        p_wb_v = load_workbook(io.BytesIO(p_bytes), data_only=True, read_only=True)
        
        check_map = {}
        for sn in p_wb_f.sheetnames:
            cells = [c.coordinate for row in p_wb_f[sn].iter_rows() for c in row if c.fill and c.fill.fill_type == 'solid' and str(c.fill.start_color.rgb)[-6:] == target_hex]
            if cells: check_map[sn] = cells
        
        total_qs = sum(len(v) for v in check_map.values())
        summary_results = []
        all_wrongs = []
        uid_re = re.compile(r'[uU]\d{7}')

        progress_bar = st.progress(0)
        for i, s_file in enumerate(student_files):
            with st.status(f"Grading {s_file.name}...", expanded=True) as status:
                s_bytes = s_file.read()
                with zipfile.ZipFile(io.BytesIO(s_bytes)) as z:
                    s_cache = {f: z.read(f).decode('utf-8', errors='replace') for f in z.namelist() if 'xl/worksheets/sheet' in f}
                
                s_wb_f = load_workbook(io.BytesIO(s_bytes), data_only=False, read_only=True)
                s_wb_v = load_workbook(io.BytesIO(s_bytes), data_only=True, read_only=True)
                uid = uid_re.search(s_file.name).group() if uid_re.search(s_file.name) else s_file.name

                correct = 0
                for sn, cells in check_map.items():
                    if sn not in s_wb_f.sheetnames: continue
                    for c in cells:
                        is_sl, s_ans, p_ans, msg = check_sparkline_advanced(p_cache, s_cache, c)
                        if is_sl == "Skip":
                            pf, sf = p_wb_f[sn][c].value, s_wb_f[sn][c].value
                            pv, sv = p_wb_v[sn][c].value, s_wb_v[sn][c].value
                            if check_logic_equivalence(pf, sf, pv, sv):
                                correct += 1
                            else:
                                # Store info without AI call
                                all_wrongs.append({"UnID": uid, "Sheet": sn, "Cell": c, "Stud_Ans": format_ans(sf, sv), "Prof_Ans": format_ans(pf, pv)})
                        else: 
                            if is_sl:
                                correct += 1
                            else:
                                all_wrongs.append({"UnID": uid, "Sheet": sn, "Cell": c, "Stud_Ans": f"[Sparkline] {s_ans}", "Prof_Ans": f"[Sparkline] {p_ans}"})
                status.update(label=f" {uid} Done!", state="complete", expanded=False)
            summary_results.append({"UnID": uid, "Score": f"{correct}/{total_qs}", "Raw": correct})
            progress_bar.progress((i + 1) / len(student_files))

        st.session_state.update({'summary_df': pd.DataFrame(summary_results).sort_values("Raw", ascending=False), 'all_wrongs_list': all_wrongs, 'grading_done': True, 'total_questions': total_qs})
        st.rerun()

if st.session_state['grading_done']:
    df_all_errors = pd.DataFrame(st.session_state['all_wrongs_list'])

    st.divider()
    col_chart, col_table = st.columns([1, 1])
    with col_chart:
        st.subheader("📊 Class Error Analysis")
        if not df_all_errors.empty:
            st.bar_chart(df_all_errors['Cell'].value_counts())
        else:
            st.success("No errors detected! 🎉")
    with col_table:
        st.subheader("📋 Summary")
        st.dataframe(st.session_state['summary_df'].drop(columns=["Raw"]), hide_index=True)

    st.divider()
    st.subheader("📥 Download Final Reports")
    # Split into two columns for Excel and PDF sections
    col_dl1, col_dl2 = st.columns(2)
    
    with col_dl1:
        xlsx_report = io.BytesIO()
        with pd.ExcelWriter(xlsx_report, engine='openpyxl') as writer:
            st.session_state['summary_df'].drop(columns=['Raw']).to_excel(writer, index=False, sheet_name="Summary")
            if not df_all_errors.empty:
                df_all_errors.to_excel(writer, index=False, sheet_name="All_Errors")
        st.download_button("📊 Download Excel Summary", xlsx_report.getvalue(), "IS 2010 Results.xlsx", width='stretch')

    with col_dl2:
        # Step 1: Button to generate AI reports
        if st.button("📄 Step 1: Generate AI Reports", width='stretch'):
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zf:
                for uid in st.session_state['summary_df']['UnID'].tolist():
                    with st.spinner(f"Generating AI Feedback for {uid}..."):
                        s_errs = [e for e in st.session_state['all_wrongs_list'] if e['UnID'] == uid]
                        s_score = st.session_state['summary_df'][st.session_state['summary_df']['UnID'] == uid]['Score'].values[0]
                        pdf_data = create_pdf_report(uid, s_score, s_errs)
                        zf.writestr(f"Report_{uid}.pdf", pdf_data)
            
            st.session_state['zip_data'] = zip_buffer.getvalue()
            st.success("✅ All reports generated!")

        # Step 2: Download button appears once ZIP is ready
        if 'zip_data' in st.session_state:
            st.write("")
            # Center the button using 1:2:1 column ratio
            _, center_col, _ = st.columns([1, 2, 1])
            with center_col:
                st.download_button(
                    label="📥 Click to Save My ZIP",
                    data=st.session_state['zip_data'],
                    file_name="Student_Reports.zip",
                    mime="application/zip",
                    width='stretch'
                )
else:
    st.info("Upload files to start.")
